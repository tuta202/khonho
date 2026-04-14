import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload

from app.dependencies import get_current_user, get_db, require_owner
from app.models.inventory import Inventory
from app.models.product import Product
from app.models.transaction import Transaction
from app.models.variant import Variant
from app.models.warehouse import Warehouse
from app.schemas.report import (
    InventorySnapshotItem,
    InventorySnapshotResponse,
    InventoryValueResponse,
    TopProductItem,
    TransactionsSummaryResponse,
    WarehouseValueItem,
)

router = APIRouter()


def _fmt_display_name(attrs) -> str:
    if not attrs:
        return "Mac dinh"
    return " / ".join(a.attr_value for a in attrs)


# ---------------------------------------------------------------------------
# Excel / CSV helpers
# ---------------------------------------------------------------------------

def _create_excel(
    headers: list[str],
    rows: list[list],
    sheet_name: str,
    bold_last_row: bool = False,
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    if bold_last_row and rows:
        last_row_idx = len(rows) + 1
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=last_row_idx, column=col_idx).font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _create_csv(headers: list[str], rows: list[list]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


def _xlsx_response(buffer: io.BytesIO, filename: str) -> StreamingResponse:
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


def _csv_response(content: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
    )


# ---------------------------------------------------------------------------
# Inventory snapshot
# ---------------------------------------------------------------------------

@router.get("/inventory-snapshot", response_model=InventorySnapshotResponse)
def inventory_snapshot(
    warehouse_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = (
        db.query(Inventory, Variant, Product, Warehouse)
        .join(Variant, Inventory.variant_id == Variant.id)
        .join(Product, Variant.product_id == Product.id)
        .join(Warehouse, Inventory.warehouse_id == Warehouse.id)
        .options(selectinload(Variant.attributes))
        .filter(
            Variant.is_active == True,    # noqa: E712
            Product.is_active == True,    # noqa: E712
            Warehouse.is_active == True,  # noqa: E712
        )
        .order_by(Product.name, Variant.id, Warehouse.name)
    )

    if warehouse_id:
        query = query.filter(Inventory.warehouse_id == warehouse_id)

    rows = query.all()

    items = []
    for inv, variant, product, warehouse in rows:
        is_low = inv.quantity <= product.low_stock_threshold
        items.append(
            InventorySnapshotItem(
                product_id=product.id,
                product_name=product.name,
                sku=product.sku,
                category=product.category,
                variant_id=variant.id,
                display_name=_fmt_display_name(variant.attributes),
                sku_variant=variant.sku_variant,
                warehouse_id=warehouse.id,
                warehouse_name=warehouse.name,
                quantity=inv.quantity,
                low_stock_threshold=product.low_stock_threshold,
                is_low_stock=is_low,
            )
        )

    return InventorySnapshotResponse(
        items=items,
        total_items=len(items),
        low_stock_count=sum(1 for i in items if i.is_low_stock),
    )


@router.get("/inventory-snapshot/export")
def export_inventory_snapshot(
    format: str = Query("xlsx"),
    warehouse_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = (
        db.query(Inventory, Variant, Product, Warehouse)
        .join(Variant, Inventory.variant_id == Variant.id)
        .join(Product, Variant.product_id == Product.id)
        .join(Warehouse, Inventory.warehouse_id == Warehouse.id)
        .options(selectinload(Variant.attributes))
        .filter(
            Variant.is_active == True,    # noqa: E712
            Product.is_active == True,    # noqa: E712
            Warehouse.is_active == True,  # noqa: E712
        )
        .order_by(Product.name, Variant.id, Warehouse.name)
    )
    if warehouse_id:
        query = query.filter(Inventory.warehouse_id == warehouse_id)

    rows = query.all()

    headers = ["San pham", "Bien the", "Kho", "So luong", "Gia von", "Gia ban", "Trang thai"]
    data_rows = []
    for inv, variant, product, warehouse in rows:
        cost = float(variant.cost_price_override if variant.cost_price_override is not None else product.cost_price)
        status = "Sap het" if inv.quantity <= product.low_stock_threshold else "Binh thuong"
        data_rows.append([
            product.name,
            _fmt_display_name(variant.attributes),
            warehouse.name,
            inv.quantity,
            cost,
            float(product.selling_price),
            status,
        ])

    filename = f"ton-kho-{date.today().isoformat()}"

    if format == "csv":
        return _csv_response(_create_csv(headers, data_rows), filename)
    return _xlsx_response(_create_excel(headers, data_rows, "Ton kho"), filename)


# ---------------------------------------------------------------------------
# Transactions summary
# ---------------------------------------------------------------------------

_TX_TYPE_LABELS = {
    "import": "Nhap hang",
    "export_sale": "Xuat ban",
    "export_damage": "Xuat huy",
    "transfer": "Chuyen kho",
    "adjust": "Dieu chinh",
}


@router.get("/transactions-summary", response_model=TransactionsSummaryResponse)
def transactions_summary(
    date_from: str = Query(...),
    date_to: str = Query(...),
    warehouse_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    try:
        dt_from = datetime.fromisoformat(date_from)
        dt_to = datetime.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="date_from và date_to phải là ISO date string (YYYY-MM-DD)",
        )

    base = (
        db.query(Transaction)
        .filter(
            Transaction.created_at >= dt_from,
            Transaction.created_at <= dt_to,
        )
    )
    if warehouse_id:
        base = base.filter(
            (Transaction.from_warehouse_id == warehouse_id)
            | (Transaction.to_warehouse_id == warehouse_id)
        )

    txs = base.all()

    by_type: dict[str, int] = {}
    for tx in txs:
        by_type[tx.type] = by_type.get(tx.type, 0) + tx.quantity

    total_imported = by_type.get("import", 0)
    total_exported = (
        by_type.get("export_sale", 0)
        + by_type.get("export_damage", 0)
        + by_type.get("adjust", 0)
    )

    product_totals: dict[int, dict] = {}
    for tx in txs:
        variant = db.query(Variant).filter(Variant.id == tx.variant_id).first()
        if not variant:
            continue
        product = db.query(Product).filter(Product.id == variant.product_id).first()
        if not product:
            continue
        pid = product.id
        if pid not in product_totals:
            product_totals[pid] = {
                "product_id": pid,
                "product_name": product.name,
                "total_imported": 0,
                "total_exported": 0,
            }
        if tx.type == "import":
            product_totals[pid]["total_imported"] += tx.quantity
        elif tx.type in ("export_sale", "export_damage", "adjust"):
            product_totals[pid]["total_exported"] += tx.quantity

    top_products = sorted(
        product_totals.values(),
        key=lambda x: x["total_imported"],
        reverse=True,
    )[:10]

    return TransactionsSummaryResponse(
        period={"from": date_from, "to": date_to},
        total_imported=total_imported,
        total_exported=total_exported,
        by_type=by_type,
        top_products=[TopProductItem(**p) for p in top_products],
    )


@router.get("/transactions-summary/export")
def export_transactions_summary(
    format: str = Query("xlsx"),
    date_from: str = Query(...),
    date_to: str = Query(...),
    warehouse_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    try:
        dt_from = datetime.fromisoformat(date_from)
        dt_to = datetime.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="date_from và date_to phải là ISO date string (YYYY-MM-DD)",
        )

    query = (
        db.query(Transaction)
        .options(
            selectinload(Transaction.variant).selectinload(Variant.product),
            selectinload(Transaction.variant).selectinload(Variant.attributes),
            selectinload(Transaction.from_warehouse),
            selectinload(Transaction.to_warehouse),
            selectinload(Transaction.user),
        )
        .filter(
            Transaction.created_at >= dt_from,
            Transaction.created_at <= dt_to,
        )
        .order_by(Transaction.created_at.desc())
    )
    if warehouse_id:
        query = query.filter(
            (Transaction.from_warehouse_id == warehouse_id)
            | (Transaction.to_warehouse_id == warehouse_id)
        )

    txs = query.all()

    headers = [
        "Thoi gian", "Loai giao dich", "San pham", "Bien the",
        "So luong", "Kho nguon", "Kho dich", "Nguoi thuc hien", "Ghi chu",
    ]
    data_rows = []
    for tx in txs:
        data_rows.append([
            tx.created_at.strftime("%d/%m/%Y %H:%M"),
            _TX_TYPE_LABELS.get(tx.type, tx.type),
            tx.variant.product.name if tx.variant and tx.variant.product else "",
            _fmt_display_name(tx.variant.attributes) if tx.variant else "",
            tx.quantity,
            tx.from_warehouse.name if tx.from_warehouse else "",
            tx.to_warehouse.name if tx.to_warehouse else "",
            tx.user.name if tx.user else "",
            tx.note or "",
        ])

    filename = f"giao-dich-{date_from}-{date_to}"

    if format == "csv":
        return _csv_response(_create_csv(headers, data_rows), filename)
    return _xlsx_response(_create_excel(headers, data_rows, "Giao dich"), filename)


# ---------------------------------------------------------------------------
# Inventory value — owner only
# ---------------------------------------------------------------------------

@router.get("/inventory-value", response_model=InventoryValueResponse)
def inventory_value(
    db: Session = Depends(get_db),
    _=Depends(require_owner),
):
    effective_cost = func.coalesce(Variant.cost_price_override, Product.cost_price)

    rows = (
        db.query(
            Warehouse.id,
            Warehouse.name,
            func.sum(Inventory.quantity * effective_cost).label("total_value"),
            func.sum(Inventory.quantity).label("total_quantity"),
            func.count(Inventory.id).label("variant_count"),
        )
        .join(Inventory, Warehouse.id == Inventory.warehouse_id)
        .join(Variant, Inventory.variant_id == Variant.id)
        .join(Product, Variant.product_id == Product.id)
        .filter(
            Warehouse.is_active == True,  # noqa: E712
            Variant.is_active == True,    # noqa: E712
            Product.is_active == True,    # noqa: E712
        )
        .group_by(Warehouse.id, Warehouse.name)
        .order_by(Warehouse.name)
        .all()
    )

    items = [
        WarehouseValueItem(
            warehouse_id=row.id,
            warehouse_name=row.name,
            total_value=Decimal(str(row.total_value or 0)),
            total_quantity=int(row.total_quantity or 0),
            variant_count=int(row.variant_count or 0),
        )
        for row in rows
    ]

    grand_total_value = sum((i.total_value for i in items), Decimal("0"))
    grand_total_quantity = sum(i.total_quantity for i in items)

    return InventoryValueResponse(
        items=items,
        grand_total_value=grand_total_value,
        grand_total_quantity=grand_total_quantity,
    )


@router.get("/inventory-value/export")
def export_inventory_value(
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    _=Depends(require_owner),
):
    rows_data = (
        db.query(Inventory, Variant, Product, Warehouse)
        .join(Variant, Inventory.variant_id == Variant.id)
        .join(Product, Variant.product_id == Product.id)
        .join(Warehouse, Inventory.warehouse_id == Warehouse.id)
        .options(selectinload(Variant.attributes))
        .filter(
            Warehouse.is_active == True,  # noqa: E712
            Variant.is_active == True,    # noqa: E712
            Product.is_active == True,    # noqa: E712
        )
        .order_by(Warehouse.name, Product.name)
        .all()
    )

    headers = ["Kho", "San pham", "Bien the", "So luong", "Gia von", "Gia tri"]
    data_rows = []
    grand_total = Decimal("0")

    for inv, variant, product, warehouse in rows_data:
        cost = variant.cost_price_override if variant.cost_price_override is not None else product.cost_price
        value = Decimal(str(inv.quantity)) * cost
        grand_total += value
        data_rows.append([
            warehouse.name,
            product.name,
            _fmt_display_name(variant.attributes),
            inv.quantity,
            float(cost),
            float(value),
        ])

    data_rows.append(["TONG", "", "", "", "", float(grand_total)])

    filename = f"gia-tri-ton-kho-{date.today().isoformat()}"

    if format == "csv":
        return _csv_response(_create_csv(headers, data_rows), filename)
    return _xlsx_response(
        _create_excel(headers, data_rows, "Gia tri ton kho", bold_last_row=True),
        filename,
    )
